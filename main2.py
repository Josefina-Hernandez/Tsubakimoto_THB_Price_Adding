import openpyxl as xl

print('Reading source file...')
wb = xl.load_workbook(filename='Output.xlsx', data_only=True)
ws = wb['NOV 2023']

source_models = []
for i in range(5, ws.max_row+1):
    source_models.append([str(ws.cell(row=i, column=3).value).strip(), str(ws.cell(row=i, column=4).value).strip()])

wb.close()


wb = xl.load_workbook(filename='Source_data.xlsx', data_only=True)
ws = wb.active

total = ws.max_row
for i in range(2, ws.max_row+1):
    print(f'Processing line: {i}/{total}...')
    ws.cell(row=i, column=5).value = 'No'
    for each_model in source_models:
        if str(ws.cell(row=i, column=2).value).strip() == each_model[0] or str(ws.cell(row=i, column=2).value).strip() == each_model[1]:
            ws.cell(row=i, column=5).value = 'Yes'
            break

wb.save(filename='Source_data.xlsx')
wb.close()

print('Finished!')
import openpyxl as xl

#wb = xl.load_workbook(filename="./2.Distributors Sprocket price list (Fix price THB) (March 2024-Rev. Nichiden Mul).xlsx", data_only=True)

wb = xl.load_workbook(filename="./3.THB Small size conveyor chain  (March 2024-Rev. Nichiden Mul).xlsx", data_only=True)

sheet_names = wb.sheetnames

for each in sheet_names:
    print(each)

count = len(sheet_names)

print(count)

output_data = []

#for i in range(7, count):    #for file "Distributors"
for i in range(5, count):    #for file "Small size conveyor"
    ws = wb[sheet_names[i]]

    print(sheet_names[i])
    print(f'max row: {ws.max_row}, max column: {ws.max_column}')

    model1_header_position = []
    unit1_header_position = []
    price1_header_position = []
    model2_header_position = []
    unit2_header_position = []
    price2_header_position = []
    model_detected = False
    unit_detected = False
    price_detected = False
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            if str(ws.cell(row=i, column=j).value).strip().lower() == "model":
                if not model_detected:
                    model1_header_position = [i, j]
                    model_detected = True
                else:
                    model2_header_position = [i, j]
            elif str(ws.cell(row=i, column=j).value).strip().lower() == "unit":
                if not unit_detected:
                    unit1_header_position = [i, j]
                    unit_detected = True
                else:
                    unit2_header_position = [i, j]
            elif str(ws.cell(row=i, column=j).value).strip().lower() == "standard price (thb)" or str(ws.cell(row=i, column=j).value).strip().lower() == "standatd price (thb)":
                if not price_detected:
                    price1_header_position = [i, j]
                    price_detected = True
                else:
                    price2_header_position = [i, j]
                    if model2_header_position == [] and unit2_header_position == []:
                        price1_header_position = price2_header_position
                        price2_header_position = []

    print(f'Model1: {model1_header_position}, Unit1: {unit1_header_position}, Price1: {price1_header_position}')
    print(f'Model2: {model2_header_position}, Unit2: {unit2_header_position}, Price2: {price2_header_position}')

    previous_price = ''
    for i in range(model1_header_position[0]+1, ws.max_row + 1):
        model1 = ws.cell(row = i, column = model1_header_position[1]).value
        unit1 = ws.cell(row = i, column = unit1_header_position[1]).value
        price1 = ws.cell(row = i, column = price1_header_position[1]).value

        if not model1 or not unit1:
            continue

        else:
            if model1 and not price1:
                price1 = previous_price
            output_data.append([str(model1).strip(), str(unit1).strip(), price1])
            previous_price = price1


    if model2_header_position:
        previous_price = ''
        for i in range(model2_header_position[0] + 1, ws.max_row + 1):
            model2 = ws.cell(row=i, column=model2_header_position[1]).value
            unit2 = ws.cell(row=i, column=unit2_header_position[1]).value
            price2 = ws.cell(row=i, column=price2_header_position[1]).value

            if not model2 or not unit2:
                continue

            else:
                if model2 and not price2:
                    price2 = previous_price
                output_data.append([str(model2).strip(), str(unit2).strip(), price2])
                previous_price = price2

counter = 0
for each in output_data:
    counter += 1
    print(counter, each)

wb.close()